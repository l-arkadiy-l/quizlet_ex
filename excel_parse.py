import openpyxl as ox
from translator import *


def update_spreadsheet(path: str, ws_len: int, do_mark: [int, bool], list_remove_words: list) -> None:
    wb = ox.load_workbook(path)
    sheet = wb.active
    if do_mark:
        for i in list_remove_words:
            sheet[f'B{i}'] = 1
    for i in range(1, ws_len + 1):
        sheet[f'C{i}'] = f'=IF(B{i}<>1,A{i},)'
    wb.save(path)


def go(file, quantity_of_words: int):
    quantity_of_words = abs(quantity_of_words)
    eng_file = rf'{file}'
    wb = ox.load_workbook(eng_file)
    sheet = wb.active
    ws_len = wb[sheet.title].max_row
    print(ws_len)
    words = {i: sheet[f'A{i}'].value for i in range(1, ws_len) if sheet[f'B{i}'].value != 1}
    update_spreadsheet(eng_file, ws_len, 1, list(words.keys())[:quantity_of_words])
    return get_translate(list(words.values())[:quantity_of_words])
