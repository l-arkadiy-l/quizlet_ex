import os

from aiogram.types import ReplyKeyboardMarkup, ReplyKeyboardRemove, KeyboardButton, ParseMode, Message, \
    InlineKeyboardMarkup, InlineKeyboardButton
from aiogram import Bot, Dispatcher, executor, types
from aiogram.types import ReplyKeyboardMarkup, ReplyKeyboardRemove, KeyboardButton
import openpyxl
import sqlite3
from excel_parse import go
from translator import *

TOKEN = "Your_token"
bot = Bot(token=TOKEN)
dp = Dispatcher(bot)

keyboard_crypto = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)


def get_users_file_quantity(user_id):
    conn = sqlite3.connect('users.db')
    c = conn.cursor()
    # c.execute("CREATE TABLE users (id integer UNIQUE, path text)")
    try:
        c.execute(f"INSERT INTO users VALUES ({user_id}, 'pathxl/{user_id}.xlsx', 5, 'eng')")
        conn.commit()
    except sqlite3.IntegrityError:
        pass
    c.execute(f"SELECT * FROM users WHERE id={user_id}")
    p = c.fetchone()
    file = p[-3]
    quantity = p[-2]
    language = p[-1]
    c.close()
    return file, quantity, language


def change_language(user_id):
    conn = sqlite3.connect('users.db')
    c = conn.cursor()
    c.execute(f"SELECT * FROM users WHERE id={user_id}")
    p = c.fetchone()
    cur_language = 'rus' if p[-1] == 'eng' else 'eng'
    c.execute(f"UPDATE users set language = '{cur_language}' where id = {user_id}")
    conn.commit()
    c.close()


def change_quantity(user_id, quantity):
    conn = sqlite3.connect('users.db')
    c = conn.cursor()
    c.execute(f"UPDATE users set quantity = {quantity} where id = {user_id}")
    conn.commit()
    c.close()


def create_file(message, file: str) -> bool:
    try:
        f = open(file)
    except FileNotFoundError:
        try:
            os.remove(f'pathxl/{message.from_user.id}.xlsx')
        except Exception:
            pass
        try:
            os.rename(f'pathxl/{message.from_user.id}00.xlsx', file)
        except Exception:
            pass
        filepath = f"pathxl/{message.from_user.id}.xlsx"
        wb = openpyxl.Workbook()
        wb.save(filepath)
    return True


# /info - подробный план как пользоватья (сделать в качестве картинок)
@dp.message_handler(commands=['start'])
async def main(message):
    user_id = message.from_user.id
    user_file, quantity, language = get_users_file_quantity(user_id)
    if language == 'eng':
        await bot.send_message(message.chat.id,
                               'Hi, I will help you import words with translation to Quizlet or other apps! We hame commands:\n'
                               '/set your number\n'
                               # '/info\n'
                               # '/set_delimeter your delimeter (by default, it is //)\n'
                               '/add your words (example: /add cat, dog, town, light)\n'
                               '/change_language <- Измените язык, чтобы перейти на русский\n'
                               '/trans your words (example: /trans cat, dog, town, light)\n'
                               '/get_file\n'
                               '/import_words\n'
                               'send your xlsx file to bot')
    else:
        await bot.send_message(message.chat.id,
                               'Привет, я помогу вам импортировать слова с переводом в Quizlet или другие приложения! У нас есть команды:\n'
                               '/set число\n'
                               # '/info\n'
                               # '/set_delimeter your delimeter (by default, it is //)\n'
                               '/add твои слова (например: /add cat, dog, town, light)\n'
                               '/change_language <- Click to change language to english\n'
                               '/trans твои слова (например: /trans cat, dog, town, light)\n'
                               '/get_file\n'
                               '/import_words\n'
                               'можешь отправить xlsx файл c твоими словами боту')


@dp.message_handler(commands=['set'])
async def user_set(message):
    user_id = message.from_user.id
    user_file, quantity, language = get_users_file_quantity(user_id)
    try:
        quantity_of_words = int(message.text.split()[-1])
        change_quantity(message.from_user.id, quantity_of_words)
        await message.reply(
            f"Quantity of words wich will be print = {quantity_of_words}" if language == 'eng' else f"Количество слов, которые будут выведены = {quantity_of_words}")
        await message.reply('Ok, now click /import_words' if language == 'eng' else 'Теперь кликни на /import_words')
    except ValueError:
        await message.reply(
            f"Error -> Enter one space and one number! Example:\n/set 6" if language == 'eng' else 'Ошибка -> нажми 1 пробел и 1 число! Например:\n/set 6')


@dp.message_handler(commands=['add'])
async def user_add(message):
    user_id = message.from_user.id
    user_file, quantity, language = get_users_file_quantity(user_id)
    words = ''.join(message.text.split()[1:]).split(',')
    file_user = f'pathxl/{message.from_user.id}.xlsx'
    create_file(message, file_user)
    eng_file = rf'{file_user}'
    wb = openpyxl.load_workbook(eng_file)
    sheet = wb.active
    # O(n) -> n - count of our words in A column
    last_coord = 0
    list_empty_coords = []
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1):
        for cell in row:
            if cell.value is None:
                print(str(cell.row))
                list_empty_coords.append(cell.row)
            last_coord = cell.coordinate[1:]
    n = len(words)
    for i in range(min(len(list_empty_coords), n)):
        sheet[f'A{list_empty_coords[i]}'] = words[0]
        words.pop(0)
    last_coord = int(last_coord) + 1
    n = len(words)
    for i in range(last_coord, last_coord + n):
        sheet[f'A{i}'] = words[0]
        words.pop(0)
    wb.save(file_user)
    await bot.send_message(message.from_user.id,
                           "Your words have added" if language == 'eng' else 'Твои слова были добавлены!')


@dp.message_handler(content_types=types.ContentTypes.DOCUMENT)
async def doc_handler(message: Message):
    """func to download xlsx files"""
    # TODO: сделать параметр, чтобы пользователь мог выбироать добвлять ли ему повторяющиеся слова или нет
    user_id = message.from_user.id
    user_file, quantity, language = get_users_file_quantity(user_id)
    file_user = f'pathxl/{message.from_user.id}.xlsx'
    try:
        f = open(file_user)
        if language == 'eng':
            kb = [
                [
                    types.KeyboardButton(text="YES"),
                    types.KeyboardButton(text="NO"),
                    types.KeyboardButton(text='Show my current file')
                ],
            ]
        else:
            kb = [
                [
                    types.KeyboardButton(text="ДА"),
                    types.KeyboardButton(text="НЕТ"),
                    types.KeyboardButton(text='Покажи мой текущий файл')
                ],
            ]
        keyboard = types.ReplyKeyboardMarkup(
            keyboard=kb,
            resize_keyboard=True,
            input_field_placeholder="Выберите способ подачи",
            one_time_keyboard=True
        )
        await message.answer("Change your file?", reply_markup=keyboard)
        document = message.document
        await document.download(destination_file=f'pathxl/{message.from_user.id}00.xlsx')
    except FileNotFoundError:
        document = message.document
        await document.download(destination_file=f'pathxl/{message.from_user.id}00.xlsx')
        await message.reply("File has been upload!" if language == 'eng' else 'Файл был загружен!')


@dp.message_handler(commands=['import_words'])
async def write_in_file(message: Message):
    file_user = f'pathxl/{message.from_user.id}.xlsx'
    user_id = message.from_user.id
    create_file(message, file_user)
    user_file, quantity, language = get_users_file_quantity(user_id)
    added_words, error_words = go(file_user, quantity)
    final_text = ''
    print(added_words, error_words)
    for key, item in added_words.items():
        final_text += key + '//' + item + '\n'
    await bot.send_message(message.from_user.id,
                           final_text if final_text else (
                               "You should to add new words! Use /add command" if language == 'eng' else 'Вы должны добавить еще слов! Используйте комманду /add'))


@dp.message_handler(commands=['get_file'])
async def get_file(message: Message):
    file_user = f'pathxl/{message.from_user.id}.xlsx'
    try:
        f = open(f'pathxl/{message.from_user.id}00.xlsx')
        try:
            os.remove(f'pathxl/{message.from_user.id}.xlsx')
        except Exception:
            pass
        try:
            os.rename(f'pathxl/{message.from_user.id}00.xlsx', file_user)
        except Exception:
            pass
    except FileNotFoundError:
        pass
    create_file(message, file_user)
    await message.reply_document(open(f'pathxl/{message.from_user.id}.xlsx', 'rb'))


@dp.message_handler(commands=['trans'])
async def write_in_file(message: Message):
    text = [i.strip() for i in ' '.join(message.text.split(' ')[1:]).split(',')]
    final_text = ''
    c = 0
    added_words, error_words = get_translate(text, is_save=False)
    for key, item in added_words.items():
        c += 1
        final_text += str(c) + '. ' + item + '\n'
    if added_words:
        await bot.send_message(message.from_user.id, final_text)


@dp.message_handler(commands=['change_language'])
async def command_change_language(message: Message):
    user_id = message.from_user.id
    change_language(user_id)
    user_file, quantity, language = get_users_file_quantity(user_id)
    await bot.send_message(user_id, 'Language changed successfully' if language == 'eng' else 'Язык успешно изменен')



@dp.message_handler(lambda message: message.text in ['YES', 'ДА'])
async def user_input(message):
    # check if '...00' is exist if true -> delete current file and rename new
    user_id = message.from_user.id
    user_file, quantity, language = get_users_file_quantity(user_id)
    try:
        f = open(f'pathxl/{message.from_user.id}00.xlsx')
        try:
            os.remove(f'pathxl/{message.from_user.id}.xlsx')
        except FileNotFoundError:
            pass
        f.close()
        os.rename(f'pathxl/{message.from_user.id}00.xlsx', f'pathxl/{message.from_user.id}.xlsx')
        await bot.send_message(message.from_user.id,
                               "OK, your file has been replaced with a new one" if language == 'eng' else 'Хорошо, ваш файл был заменен на новый')
    except (FileExistsError, FileNotFoundError):
        # accidentally write YES
        print('here')
        pass


@dp.message_handler(lambda message: message.text in ['NO', 'НЕТ'])
async def user_input(message):
    await bot.send_message(message.from_user.id, "Ok")


@dp.message_handler(lambda message: message.text in ['Show my current file', 'Покажи мой текущий файл'])
async def user_input(message):
    await message.reply_document(open(f'pathxl/{message.from_user.id}.xlsx', 'rb'))


@dp.message_handler(content_types=['text'])
async def user_input(message):
    user_id = message.from_user.id
    user_file, quantity, language = get_users_file_quantity(user_id)
    await bot.send_message(message.from_user.id,
                           "Sorry but I can't to understand you, try to enter other commands" if language == 'eng' else 'Извините, но я вас не понимаю! Используйте команды')


executor.start_polling(dp)
